import os

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment


def excel_read(path: str, title_num: int = 4):
    """

    :param path: 文件完整路径
    :param title_num: 标题数量
    :return: 包含标题与数据的列表
    """
    if not os.path.exists(path):
        return False
    title_rows = []
    data = []
    workbook = load_workbook(path)
    sheet = workbook.active

    for i,row in enumerate(sheet.rows):
        values=[]
        for cell in row:
            value=cell.value
            if value is None:
                value=''
            values.append(value)
        values.pop(2)
        if i < title_num:
            title_rows.append(values)
        else:
            data.append(values)
    return [title_rows, data]


def excel_save(root_path: str, title_rows: list, data_rows: list):
    """

    :param root_path: 储存生成文件的根路径
    :param title_rows: 包含标题的行
    :param data_rows: 包含数据的行
    :return:
    """
    if not os.path.exists(root_path):
        os.mkdir(root_path)

    title_num = len(title_rows)
    align = Alignment(horizontal='center', vertical='center')
    for index, data_row in enumerate(data_rows):
        dest_filename=os.path.join(root_path,data_row[0]+'.xlsx')
        wb = Workbook()

        ws = wb.worksheets[0]
        ws.title = "月度汇总"
        for i, row in enumerate(title_rows):
            for j,cell in enumerate(row):
                if i==2 and (j<8 or j>10):
                    ws.merge_cells(None,i+1,j+1,i+2,j+1)
                elif i==2 and j==9:
                    ws.merge_cells(None,i+1,j+1,i+1,j+2)
                ws.cell(i+1,j+1).value=cell
                if i>1:
                    ws.cell(i+1,j+1).alignment=align

        for i in range(len(data_row)):
            ws.cell(title_num+1,i+1).value=data_row[i]
            ws.cell(title_num+1,i+1).alignment = align
        wb.save(filename=dest_filename)


if __name__ == '__main__':
    for name in os.listdir('./'):
        if not ('.xlsx' in name or 'xls' in name):
            continue
        path_dir = name.split('.xls')[0]
        if not os.path.exists('out'):
            os.mkdir('out')
        file_path = os.path.join('./out', path_dir)
        if os.path.exists(file_path):
            print(name, '已存在,跳过解析')
            continue
        else:
            os.mkdir(file_path)
        print('解析文件:', name)
        try:
            context = excel_read(name)
            excel_save(file_path, *context)
        except Exception as e:
            print('解析过程产生异常:', e)
            input()

    print('\n\n文件解析完毕,请查收')
    input()
